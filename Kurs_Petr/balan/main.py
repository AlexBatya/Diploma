from openpyxl import Workbook
import matplotlib.pyplot as plt
import math 

#передн€€ предельна€ центровка дл€ H=0,M=0.2 (посадка)
def XTTP(L_go,S_go,X_pered):               
    K_go=0.85
    mzo_bgo=-0.1425 #может быть на пор€док меньше
    xf_bgo=0.245
    alfa=6
    Cyo_bgo=0.8
    Cy_poalfa_bgo=5.4
    Cy_poalfa_go=5.2
    E_poalfa=0.57
    delta_max=-25
    fi_yct=-4
    nb=0.34
    fi_effekt=fi_yct+nb*delta_max
    Cy_bgo=Cyo_bgo+Cy_poalfa_bgo*alfa*3.14/180
    Cy_go=Cy_poalfa_go*(alfa*(1-E_poalfa)+fi_effekt)*3.14/180
    x_ttp=(-mzo_bgo+xf_bgo*Cy_bgo+Cy_go*S_go*K_go*L_go)/Cy_bgo
    X_pered.append(x_ttp)
    print("S_go=",S_go)
    print("x_ttp=",x_ttp)
    print()

#задн€€ предельна€ центровка дл€ режима H=0,M=0.3
def XTPZ(L_go,ba,g,Ps,Ro,S_go,X_zad):
    sigma_min=-0.1
    u=(2*Ps*10)/(Ro*g*ba)
    K_go=0.85
    Cy_poalfa_go=5.2
    Cy_poalfa=5.5
    E_poalfa=0.57
    xF_bgo=0.245
    mz_powz_bgo=-3.1
    mz_powz_go=-Cy_poalfa_go*S_go*L_go**2*math.sqrt(K_go)
    delta_xF=(Cy_poalfa_go/Cy_poalfa)*(1-E_poalfa)*S_go*L_go*K_go
    xF=xF_bgo+delta_xF
    mz_powz=mz_powz_bgo+mz_powz_go
    xH=xF-mz_powz/u
    x_tpz=xH+sigma_min
    X_zad.append(x_tpz)
    print("S_go=",S_go)
    print('x_tpz=',x_tpz)
    print()

S_go=[0.0865,0.2]
X_pered=[]
X_zad=[]
L_go=4
ba=6.7
g=9.81
Ro=1.225
Ps=550
XTTP(L_go,S_go[0],X_pered)
XTTP(L_go,S_go[1],X_pered)
XTPZ(L_go,ba,g,Ps,Ro,S_go[0],X_zad)
XTPZ(L_go,ba,g,Ps,Ro,S_go[1],X_zad)
dx_e=0.15
delta_x=dx_e*1.2
print("delta_x=",delta_x)

plt.plot(S_go,X_pered)
plt.plot(S_go,X_zad)
plt.grid()
plt.show()
# _________________________________________________________________________________________________________________________________________________________________
from openpyxl import Workbook
import matplotlib.pyplot as plt
import math

def grafiki(M,xF,xH,x_tpz,sigma):
    plt.plot(M,xF)
    plt.plot(M,xH)
    plt.plot(M,x_tpz)
    plt.plot(M,sigma)
    plt.grid()
    plt.show()

#исходные данные 
M=[0.3,0.4,0.5,0.6,0.7,0.8,0.9]
xF_bgo=[0.245,0.245,0.245,0.245,0.245,0.265,0.4]
Cy_poalfa_go=[5.2,5.2,5.2,5.2,5.2,5.375,5.95]
Cy_poalfa=[5.5,5.5,5.5,5.6,6.15,6.55,6.55]
E_poalfa=[0.56,0.56,0.54,0.5,0.45,0.43,0.415]
K_go=[0.85,0.85,0.85,0.85,0.85,0.85,0.83]
mz_powz_bgo=[-3.1,-3.1,-3.1,-3.1,-3.125,-3.175,-3.3]
L_go=4
ba=6.7
g=9.81
Ro=1.225
Ps=550
sigma_min=-0.1
u=2*Ps*10/(Ro*g*ba)
xT=0.246
S_go=0.0865


mz_powz_go=[]
delta_xF=[]
mz_powz=[]
xH=[]
x_tpz=[]
sigma=[]
xF=[]
for i in range(len(M)):
    mz_powz_go.append(-Cy_poalfa_go[i]*S_go*L_go**2*math.sqrt(K_go[i]))
    delta_xF.append((Cy_poalfa_go[i]/Cy_poalfa[i])*(1-E_poalfa[i])*S_go*L_go*K_go[i])
    xF.append(xF_bgo[i]+delta_xF[i])
    mz_powz.append(mz_powz_bgo[i]+mz_powz_go[i])
    xH.append(xF[i]-mz_powz[i]/u)
    x_tpz.append(xH[i]+sigma_min)
    sigma.append(xT-xF[i]+mz_powz[i]/u)
print("M",'     ','xF','    ','xH','    ','x_tpz','     ','sigma')
for i in range(len(M)):
    print(round(M[i],3),'     ',round(xF[i],3),'    ',round(xH[i],3),'    ',round(x_tpz[i],3),'    ',round(sigma[i],3))

grafiki(M,xF,xH,x_tpz,sigma)

wb=Workbook()
wb['Sheet'].title='report'
sh1=wb.active
data=[(M),(xF),(xH),(x_tpz),(sigma)]
for i in data:
    sh1.append(i)
wb.save('p.xlsx')
# ________________________________________________________________________________________________________________________________________________________________
from openpyxl import Workbook
import matplotlib.pyplot as plt
import math
import pandas as pd


def grafiki(M1,M2,M3,fi_bal1,fi_bal2,fi_bal3,fi_n1,fi_n2,fi_n3,ny_p1,ny_p2,ny_p3):
    plt.plot(M1,fi_bal1)
    plt.plot(M2,fi_bal2)
    plt.plot(M3,fi_bal3)
    plt.grid()
    plt.show()

    plt.plot(M1,fi_n1)
    plt.plot(M2,fi_n2)
    plt.plot(M3,fi_n3)
    plt.grid()
    plt.show()
    
    plt.plot(M1,ny_p1)
    plt.plot(M2,ny_p2)
    plt.plot(M3,ny_p3)
    plt.grid()
    plt.show()


def balanc(M,xF_bgo,Cy_poalfa_go,Cy_poalfa,E_poalfa,K_go,mz_powz_bgo,nb,Ro,PH,a,alfa_o,mz_obgo,fi_bal,fi_n,ny_p,V,Cy_dop,ny_dop):
    m_T=0.42
    m=1-0.5*m_T
    Ps=550
    fi_max=-25
    S_go=0.0865
    xT=0.246
    L_go=4
    ba=6.7
    sigma_min=-0.1
    g=9.81
    u=(2*Ps*10)/(Ro*g*ba)
    ny_e=3

    xF=[]
    mz_cy=[]
    mz_o=[]
    Cy_gp=[]
    mz_podelta=[]
    delta_xF=[]
    mz_powz_go=[]
    mz_powz=[]
    xH=[]
    x_tpz=[]
    sigma=[]

    for i in range(len(M)):
        V.append(M[i]*a*3.6)
        mz_podelta.append(-Cy_poalfa_go[i]*S_go*L_go*K_go[i]*nb[i])
        delta_xF.append((Cy_poalfa_go[i]*(1-E_poalfa[i])*S_go*L_go*K_go[i])/Cy_poalfa[i])
        xF.append(xF_bgo[i]+delta_xF[i])
        mz_cy.append(xT-xF[i])
        mz_o.append(mz_obgo[i]-S_go*L_go*K_go[i]*Cy_poalfa_go[i]*alfa_o[i]*(1-E_poalfa[i]))
        Cy_gp.append((10*Ps*m)/(0.7*PH*M[i]**2))
        ny_dop.append(min(ny_e,Cy_dop[i]/Cy_gp[i]))
        fi_bal.append((-(mz_o[i]+mz_cy[i]*Cy_gp[i])/(mz_podelta[i]*(1+mz_cy[i]/L_go)))*180/3.14+4/nb[i])
        mz_powz_go.append(-Cy_poalfa_go[i]*S_go*L_go**2*math.sqrt(K_go[i]))
        mz_powz.append(mz_powz_bgo[i]+mz_powz_go[i])
        xH.append(xF[i]-mz_powz[i]/u)
        x_tpz.append(xH[i]+sigma_min)
        sigma.append(xT-xF[i]+mz_powz[i]/u)
        fi_n.append((-Cy_gp[i]*sigma[i]/mz_podelta[i])*180/3.14)
        ny_p.append((fi_max-fi_bal[i])/fi_n[i])
    list1=({'M':M,'V_km':V,'fi_bal':fi_bal,'fi_n':fi_n,'ny_p':ny_p})
    l=pd.DataFrame(list1)
    print(l)

print('H=0 km')
fi_bal1=[]
fi_n1=[]
ny_p1=[]
V1=[]
ny_dop1=[]
Cya_dop1=[1.125,1.125,1.112,1.083,1.033,1.033]
M1=[0.256,0.3,0.4,0.5,0.6,0.612]
xF_bgo1=[0.245,0.245,0.245,0.245,0.245,0.245]
Cy_poalfa_go1=[5.2,5.2,5.2,5.2,5.2,5.2]
Cy_poalfa1=[5.5,5.5,5.5,5.5,5.6,5.63]
E_poalfa1=[0.56,0.56,0.56,0.54,0.5,0.48]
K_go1=[0.85,0.85,0.85,0.85,0.85,0.85]
mz_powz_bgo1=[-3.1,-3.1,-3.1,-3.1,-3.1,-3.1]
mz_obgo1=[-0.1425,-0.1425,-0.1425,-0.1425,-0.1425,-0.1425]
nb1=[0.34,0.34,0.34,0.34,0.34,0.34]
alfa_o1=[-0.01315,-0.013,-0.01245,-0.01178,-0.01108,-0.011]
PH1=101325
a1=340.29
Ro1=1.225
balanc(M1,xF_bgo1,Cy_poalfa_go1,Cy_poalfa1,E_poalfa1,K_go1,mz_powz_bgo1,nb1,Ro1,PH1,a1,alfa_o1,mz_obgo1,fi_bal1,fi_n1,ny_p1,V1,Cya_dop1,ny_dop1)

print("H=6km")
fi_bal2=[]
fi_n2=[]
ny_p2=[]
V2=[]
ny_dop2=[]
M2=[0.376,0.5,0.6,0.7,0.8,0.85]
Cya_dop2=[1.112,1.083,1.033,0.977,0.903,0.86]
xF_bgo2=[0.245,0.245,0.245,0.245,0.265,0.32]
Cy_poalfa_go2=[5.2,5.2,5.2,5.2,5.375,5.7]
Cy_poalfa2=[5.5,5.5,5.6,6.15,6.55,6.59]
E_poalfa2=[0.56,0.54,0.5,0.45,0.43,0.42]
K_go2=[0.85,0.85,0.85,0.85,0.85,0.85]
mz_powz_bgo2=[-3.1,-3.1,-3.1,-3.125,-3.175,-3.23]
mz_obgo2=[-0.1425,-0.1425,-0.1425,-0.1425,-0.1425,-0.1425]
nb2=[0.34,0.34,0.34,0.34,0.29,0.26]
alfa_o2=[-0.0124,-0.01178,-0.01108,-0.0102,-0.00924,-0.00875]
PH2=47218
a2=316.45
Ro2=0.539
balanc(M2,xF_bgo2,Cy_poalfa_go2,Cy_poalfa2,E_poalfa2,K_go2,mz_powz_bgo2,nb2,Ro2,PH2,a2,alfa_o2,mz_obgo2,fi_bal2,fi_n2,ny_p2,V2,Cya_dop2,ny_dop2)

print("H=10km")
fi_bal3=[]
fi_n3=[]
ny_p3=[]
V3=[]
ny_dop3=[]
Cya_dop3=[1.06,1.033,0.977,0.903,0.807]
M3=[0.511,0.6,0.7,0.8,0.85]
M2=[0.376,0.5,0.6,0.7,0.8,0.85]

xF_bgo3=[0.245,0.245,0.245,0.265,0.32]
Cy_poalfa_go3=[5.2,5.2,5.2,5.375,5.7]
Cy_poalfa3=[5.5,5.6,6.15,6.55,6.59]
E_poalfa3=[0.52,0.5,0.45,0.43,0.42]
K_go3=[0.85,0.85,0.85,0.85,0.85]
mz_powz_bgo3=[-3.1,-3.1,-3.125,-3.175,-3.23]
mz_obgo3=[-0.1425,-0.1425,-0.1425,-0.1425,-0.1425]
nb3=[0.34,0.34,0.34,0.29,0.26]
alfa_o3=[-0.01178,-0.01108,-0.0102,-0.00924,-0.00875]
PH3=26500
a3=299.53
Ro3=0.414
balanc(M3,xF_bgo3,Cy_poalfa_go3,Cy_poalfa3,E_poalfa3,K_go3,mz_powz_bgo3,nb3,Ro3,PH3,a3,alfa_o3,mz_obgo3,fi_bal3,fi_n3,ny_p3,V3,Cya_dop3,ny_dop3)

grafiki(M1,M2,M3,fi_bal1,fi_bal2,fi_bal3,fi_n1,fi_n2,fi_n3,ny_p1,ny_p2,ny_p3)

wb=Workbook()
wb['Sheet'].title='report'
sh1=wb.active
data=[(M1),(V1),(fi_bal1),(fi_n1),(ny_p1),(ny_dop1),(M2),(V2),(fi_bal2),(fi_n2),(ny_p2),(ny_dop2),(M3),(V3),(fi_bal3),(fi_n3),(ny_p3),(ny_dop3)]
for i in data:
    sh1.append(i)
wb.save('p.xlsx')
_________________________________________________________________________________________________________________________________________________________________
from openpyxl import Workbook
import matplotlib.pyplot as plt
import math
import pandas as pd


def grafiki(M1,M2,M3,fi_bal1,fi_bal2,fi_bal3,fi_n1,fi_n2,fi_n3,ny_p1,ny_p2,ny_p3):
    plt.plot(M1,fi_bal1)
    plt.plot(M2,fi_bal2)
    plt.plot(M3,fi_bal3)
    plt.grid()
    plt.show()

    plt.plot(M1,fi_n1)
    plt.plot(M2,fi_n2)
    plt.plot(M3,fi_n3)
    plt.grid()
    plt.show()
    
    plt.plot(M1,ny_p1)
    plt.plot(M2,ny_p2)
    plt.plot(M3,ny_p3)
    plt.grid()
    plt.show()


def balanc(M,mz_podelta,xF_bgo,Cy_poalfa_go,Cy_poalfa,E_poalfa,K_go,mz_powz_bgo,nb,Ro,PH,a,alfa_o,mz_obgo,fi_bal,fi_n,ny_p,V,Cy_dop,ny_dop):
    m_T=0.42
    m=1-0.5*m_T
    Ps=550
    fi_max=-25
    fi_ust=-4
    S_go=0.0865
    xT=0.246
    L_go=4
    ba=6.7
    sigma_min=-0.1
    g=9.81
    u=(2*Ps*10)/(Ro*g*ba)
    ny_e=3

    xF=[]
    mz_cy=[]
    mz_o=[]
    Cy_gp=[]
    delta_xF=[]
    mz_powz_go=[]
    mz_powz=[]
    xH=[]
    x_tpz=[]
    sigma=[]

    for i in range(len(M)):
        V.append(M[i]*a*3.6)
        delta_xF.append((Cy_poalfa_go[i]*(1-E_poalfa[i])*S_go*L_go*K_go[i])/Cy_poalfa[i])
        xF.append(xF_bgo[i]+delta_xF[i])
        mz_cy.append(xT-xF[i])
        mz_o.append(mz_obgo[i]-S_go*L_go*K_go[i]*Cy_poalfa_go[i]*alfa_o[i]*(1-E_poalfa[i]))
        Cy_gp.append((10*Ps*m)/(0.7*PH*M[i]**2))
        ny_dop.append(min(ny_e,Cy_dop[i]/Cy_gp[i]))
        fi_bal.append((-(mz_o[i]+mz_cy[i]*Cy_gp[i])/(mz_podelta[i]*(1+mz_cy[i]/L_go)))*180/3.14)
        mz_powz_go.append(-Cy_poalfa_go[i]*S_go*L_go**2*math.sqrt(K_go[i]))
        mz_powz.append(mz_powz_bgo[i]+mz_powz_go[i])
        xH.append(xF[i]-mz_powz[i]/u)
        x_tpz.append(xH[i]+sigma_min)
        sigma.append(xT-xF[i]+mz_powz[i]/u)
        fi_n.append((-Cy_gp[i]*sigma[i]/mz_podelta[i])*180/3.14)
        ny_p.append((fi_max-fi_bal[i]-fi_ust/nb[i])/fi_n[i])
    
    list1=({'M':M,'V_km':V,'fi_bal':fi_bal,'fi_n':fi_n,'ny_p':ny_p})
    l=pd.DataFrame(list1)
    print(l)

print('H=0 km')
fi_bal1=[]
fi_n1=[]
ny_p1=[]
V1=[]
ny_dop1=[]
Cya_dop1=[1.125,1.125,1.112,1.083,1.033,1.033]
M1=[0.256,0.3,0.4,0.5,0.6,0.612]
mz_podelta1=[-1.3,-1.3,-1.3,-1.3,-1.3,-1.3]
xF_bgo1=[0.245,0.245,0.245,0.245,0.245,0.245]
Cy_poalfa_go1=[5.2,5.2,5.2,5.2,5.2,5.2]
Cy_poalfa1=[5.5,5.5,5.5,5.5,5.6,5.63]
E_poalfa1=[0.56,0.56,0.56,0.54,0.5,0.48]
K_go1=[0.85,0.85,0.85,0.85,0.85,0.85]
mz_powz_bgo1=[-3.1,-3.1,-3.1,-3.1,-3.1,-3.1]
mz_obgo1=[-0.1425,-0.1425,-0.1425,-0.1425,-0.1425,-0.1425]
nb1=[0.34,0.34,0.34,0.34,0.34,0.34]
alfa_o1=[-0.01315,-0.013,-0.01245,-0.01178,-0.01108,-0.011]
PH1=101325
a1=340.29
Ro1=1.225
balanc(M1,mz_podelta1,xF_bgo1,Cy_poalfa_go1,Cy_poalfa1,E_poalfa1,K_go1,mz_powz_bgo1,nb1,Ro1,PH1,a1,alfa_o1,mz_obgo1,fi_bal1,fi_n1,ny_p1,V1,Cya_dop1,ny_dop1)

print("H=6km")
fi_bal2=[]
fi_n2=[]
ny_p2=[]
V2=[]
ny_dop2=[]
M2=[0.376,0.5,0.6,0.7,0.8,0.85]
mz_podelta2=[-1.3,-1.3,-1.3,-1.236,-1.2,-1.076]
Cya_dop2=[1.112,1.083,1.033,0.977,0.903,0.86]
xF_bgo2=[0.245,0.245,0.245,0.245,0.265,0.32]
Cy_poalfa_go2=[5.2,5.2,5.2,5.2,5.375,5.7]
Cy_poalfa2=[5.5,5.5,5.6,6.15,6.55,6.59]
E_poalfa2=[0.56,0.54,0.5,0.45,0.43,0.42]
K_go2=[0.85,0.85,0.85,0.85,0.85,0.85]
mz_powz_bgo2=[-3.1,-3.1,-3.1,-3.125,-3.175,-3.23]
mz_obgo2=[-0.1425,-0.1425,-0.1425,-0.1425,-0.1425,-0.1425]
nb2=[0.34,0.34,0.34,0.34,0.29,0.26]
alfa_o2=[-0.0124,-0.01178,-0.01108,-0.0102,-0.00924,-0.00875]
PH2=47218
a2=316.45
Ro2=0.539
balanc(M2,mz_podelta2,xF_bgo2,Cy_poalfa_go2,Cy_poalfa2,E_poalfa2,K_go2,mz_powz_bgo2,nb2,Ro2,PH2,a2,alfa_o2,mz_obgo2,fi_bal2,fi_n2,ny_p2,V2,Cya_dop2,ny_dop2)

print("H=10km")
fi_bal3=[]
fi_n3=[]
ny_p3=[]
V3=[]
ny_dop3=[]
Cya_dop3=[1.06,1.033,0.977,0.903,0.807]
M3=[0.511,0.6,0.7,0.8,0.85]
mz_podelta3=[-1.3,-1.3,-1.236,-1.2,-1.076]
xF_bgo3=[0.245,0.245,0.245,0.265,0.32]
Cy_poalfa_go3=[5.2,5.2,5.2,5.375,5.7]
Cy_poalfa3=[5.5,5.6,6.15,6.55,6.59]
E_poalfa3=[0.52,0.5,0.45,0.43,0.42]
K_go3=[0.85,0.85,0.85,0.85,0.85]
mz_powz_bgo3=[-3.1,-3.1,-3.125,-3.175,-3.23]
mz_obgo3=[-0.1425,-0.1425,-0.1425,-0.1425,-0.1425]
nb3=[0.34,0.34,0.34,0.29,0.26]
alfa_o3=[-0.01178,-0.01108,-0.0102,-0.00924,-0.00875]
PH3=26500
a3=299.53
Ro3=0.414
balanc(M3,mz_podelta3,xF_bgo3,Cy_poalfa_go3,Cy_poalfa3,E_poalfa3,K_go3,mz_powz_bgo3,nb3,Ro3,PH3,a3,alfa_o3,mz_obgo3,fi_bal3,fi_n3,ny_p3,V3,Cya_dop3,ny_dop3)

grafiki(M1,M2,M3,fi_bal1,fi_bal2,fi_bal3,fi_n1,fi_n2,fi_n3,ny_p1,ny_p2,ny_p3)

wb=Workbook()
wb['Sheet'].title='report'
sh1=wb.active
data=[(M1),(V1),(fi_bal1),(fi_n1),(ny_p1),(ny_dop1),(M2),(V2),(fi_bal2),(fi_n2),(ny_p2),(ny_dop2),(M3),(V3),(fi_bal3),(fi_n3),(ny_p3),(ny_dop3)]
for i in data:
    sh1.append(i)
wb.save('p.xlsx ')